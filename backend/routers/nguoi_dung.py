from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Request
from pydantic import BaseModel
from typing import Optional
import openpyxl
import io
from database import supabase
from auth import (hash_mat_khau, lay_nguoi_dung_hien_tai, chi_giao_vien,
                  chi_quan_tri, thu_hoi_phien)
import audit

router = APIRouter()

MAT_KHAU_MAC_DINH = "Okr@12345"   # [v2.6] chi con dung lam phuong an du phong

# ============================================================
#  [v2.6] SINH MAT KHAU NGAU NHIEN RIENG CHO TUNG TAI KHOAN
#  Truoc day MOI tai khoan moi deu dung chung mat khau "Okr@12345"
#  -> ai cung doan duoc, va HS chua kip doi thi tai khoan bi lo.
#  Nay moi nguoi mot mat khau rieng, chi hien DUY NHAT mot lan cho
#  giao vien/quan tri phat lai; CSDL chi luu ban bam.
# ============================================================
_CHU_HOA = "ABCDEFGHJKLMNPQRSTUVWXYZ"      # bo I, O de tranh nham voi 1, 0
_CHU_THUONG = "abcdefghijkmnpqrstuvwxyz"   # bo l
_CHU_SO = "23456789"                        # bo 0, 1


def tao_mat_khau_ngau_nhien(do_dai: int = 10) -> str:
    """Sinh mat khau de doc, de doc chinh ta, van du manh (>=1 hoa, >=1 so)."""
    import secrets
    kho = _CHU_HOA + _CHU_THUONG + _CHU_SO
    ky_tu = [
        secrets.choice(_CHU_HOA),
        secrets.choice(_CHU_SO),
        secrets.choice(_CHU_THUONG),
    ]
    ky_tu += [secrets.choice(kho) for _ in range(max(0, do_dai - len(ky_tu)))]
    secrets.SystemRandom().shuffle(ky_tu)
    return "".join(ky_tu)

class ThemHocSinh(BaseModel):
    ho_ten: str
    email: str
    email_phu_huynh: Optional[str] = None
    ten_lop: str

class ThemGiaoVien(BaseModel):
    ho_ten: str
    email: str
    ten_lop: str
    si_so: Optional[int] = None

@router.get("/hoc-sinh/{ten_lop}")
def danh_sach_hoc_sinh(ten_lop: str, nguoi_dung=Depends(chi_giao_vien)):
    # Thu exact match truoc
    res = supabase.table("nguoi_dung").select("id, ho_ten, email, email_phu_huynh, ten_lop, so_thu_tu, dang_hoat_dong, ngay_tao").eq("ten_lop", ten_lop).eq("vai_tro", "hoc_sinh").eq("dang_hoat_dong", True).order("so_thu_tu").execute()
    if res.data:
        return res.data
    # Fallback: tim khong phan biet hoa thuong va khoang trang
    res2 = supabase.table("nguoi_dung").select("id, ho_ten, email, email_phu_huynh, ten_lop, so_thu_tu, dang_hoat_dong, ngay_tao").ilike("ten_lop", ten_lop.strip()).eq("vai_tro", "hoc_sinh").eq("dang_hoat_dong", True).order("so_thu_tu").execute()
    return res2.data

@router.get("/tat-ca-hoc-sinh")
def tat_ca_hoc_sinh(nguoi_dung=Depends(chi_giao_vien)):
    """Tra ve tat ca HS (de kiem tra ten_lop luu trong DB)"""
    res = supabase.table("nguoi_dung").select("id, ho_ten, email, ten_lop, so_thu_tu, dang_hoat_dong").eq("vai_tro", "hoc_sinh").eq("dang_hoat_dong", True).order("ten_lop").execute()
    return res.data

@router.post("/hoc-sinh")
def them_hoc_sinh(body: ThemHocSinh, nguoi_dung=Depends(chi_giao_vien)):
    kiem_tra = supabase.table("nguoi_dung").select("id").eq("email", body.email).execute()
    if kiem_tra.data:
        raise HTTPException(status_code=400, detail="Email nay da ton tai trong he thong")

    data = {
        "ho_ten": body.ho_ten,
        "email": body.email,
        "email_phu_huynh": body.email_phu_huynh,
        "ten_lop": body.ten_lop,
        "vai_tro": "hoc_sinh",
        "mat_khau_hash": hash_mat_khau(MAT_KHAU_MAC_DINH),
        "bat_buoc_doi_mat_khau": True
    }
    res = supabase.table("nguoi_dung").insert(data).execute()
    return {"message": "Tao hoc sinh thanh cong", "id": res.data[0]["id"]}

@router.post("/giao-vien")
def them_giao_vien(body: ThemGiaoVien, nguoi_dung=Depends(chi_quan_tri)):
    kiem_tra = supabase.table("nguoi_dung").select("id").eq("email", body.email).execute()
    if kiem_tra.data:
        raise HTTPException(status_code=400, detail="Email nay da ton tai trong he thong")

    data = {
        "ho_ten": body.ho_ten,
        "email": body.email,
        "ten_lop": body.ten_lop,
        "si_so": body.si_so,
        "vai_tro": "giao_vien",
        "mat_khau_hash": hash_mat_khau(MAT_KHAU_MAC_DINH),
        "bat_buoc_doi_mat_khau": True
    }
    res = supabase.table("nguoi_dung").insert(data).execute()
    return {"message": "Tao giao vien thanh cong", "id": res.data[0]["id"]}

@router.post("/nhap-danh-sach")
async def nhap_danh_sach(vai_tro: str, ten_lop: Optional[str] = None, file: UploadFile = File(...), nguoi_dung=Depends(lay_nguoi_dung_hien_tai)):
    if nguoi_dung["vai_tro"] not in ["quan_tri", "giao_vien"]:
        raise HTTPException(status_code=403, detail="Khong co quyen")

    # Neu frontend khong gui ten_lop, lay tu JWT cua giao vien
    ten_lop_mac_dinh = ten_lop or nguoi_dung.get("ten_lop") or ""

    noi_dung = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(noi_dung))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel khong hop le: {str(e)}")
    ws = wb.active

    thanh_cong = 0
    loi = []
    mat_khau_da_cap = []

    # Tim cot email bang cach scan du lieu thuc te (khong doan tu header)
    # Scan toi da 5 dong dau de tim cot nao chua "@"
    email_col_idx = None
    for scan_row in ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True):
        if not scan_row: continue
        for ci, val in enumerate(scan_row):
            if val and "@" in str(val):
                email_col_idx = ci
                break
        if email_col_idx is not None:
            break

    if email_col_idx is None:
        raise HTTPException(status_code=400, detail="Khong tim thay cot email trong file. Kiem tra lai dinh dang file.")

    # offset = so cot truoc cot email (= 1 neu co STT, = 0 neu khong co STT)
    # Cot ho_ten = email_col_idx - 1
    # Cot STT    = email_col_idx - 2 (neu co)
    has_stt = email_col_idx >= 2
    offset = email_col_idx - 1  # vi tri bat dau cua ho_ten

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Bo qua dong trong hoac dong ghi chu (khong co @ o bat ky o nao)
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row_has_email = any(v and "@" in str(v) for v in row)
        if not row_has_email:
            continue
        try:
            # STT
            if has_stt and row[0] is not None:
                try:
                    stt = int(float(str(row[0]).strip()))
                except (ValueError, TypeError):
                    stt = i - 1
            else:
                stt = i - 1

            def _get(ci: int) -> str:
                return str(row[ci]).strip() if len(row) > ci and row[ci] is not None and str(row[ci]).strip() else ""

            ho_ten = _get(offset)
            email  = _get(email_col_idx)

            gioi_tinh = None
            if vai_tro == "giao_vien":
                lop       = _get(email_col_idx + 1) or ten_lop_mac_dinh
                si_so_str = _get(email_col_idx + 2)
                si_so     = int(float(si_so_str)) if si_so_str else None
                email_ph  = None
                gt_raw = _get(email_col_idx + 3).lower()
                if gt_raw in ("nam", "male", "m", "thay", "thầy"):
                    gioi_tinh = "nam"
                elif gt_raw in ("nu", "nữ", "female", "f", "co", "cô"):
                    gioi_tinh = "nu"
            else:
                email_ph = _get(email_col_idx + 1) or None
                if email_ph in ("None", ""): email_ph = None
                lop      = _get(email_col_idx + 2) or ten_lop_mac_dinh
                si_so    = None

            if not ho_ten:
                loi.append(f"Dong {i}: Thieu ho ten")
                continue
            if not email or "@" not in email:
                loi.append(f"Dong {i}: Email khong hop le: '{email}'")
                continue
            if vai_tro != "giao_vien" and not lop:
                loi.append(f"Dong {i}: Khong xac dinh duoc lop (GV chua duoc phan lop)")
                continue

            kiem_tra = supabase.table("nguoi_dung").select("id, dang_hoat_dong").eq("email", email).execute()
            if kiem_tra.data:
                cu = kiem_tra.data[0]
                if not cu.get("dang_hoat_dong", True):
                    # Email ton tai nhung dang bi vo hieu hoa -> kich hoat lai + cap nhat thong tin
                    capnhat = {
                        "ho_ten": ho_ten,
                        "ten_lop": lop or None,
                        "so_thu_tu": stt,
                        "dang_hoat_dong": True,
                    }
                    if email_ph is not None:
                        capnhat["email_phu_huynh"] = email_ph
                    supabase.table("nguoi_dung").update(capnhat).eq("id", cu["id"]).execute()
                    thanh_cong += 1
                else:
                    loi.append(f"Dong {i}: Email '{email}' da ton tai")
                continue

            record: dict = {
                "ho_ten": ho_ten,
                "email": email,
                "email_phu_huynh": email_ph,
                "ten_lop": lop or None,
                "vai_tro": vai_tro,
                "so_thu_tu": stt,
                "mat_khau_hash": hash_mat_khau(MAT_KHAU_MAC_DINH),
                "bat_buoc_doi_mat_khau": True,
                "dang_hoat_dong": True,
            }
            # [v2.6] Moi tai khoan mot mat khau ngau nhien rieng (khong dung chung)
            mat_khau_cap = tao_mat_khau_ngau_nhien()
            record["mat_khau_hash"] = hash_mat_khau(mat_khau_cap)
            if si_so is not None:
                record["si_so"] = si_so
            if gioi_tinh is not None:
                record["gioi_tinh"] = gioi_tinh
            try:
                supabase.table("nguoi_dung").insert(record).execute()
            except Exception:
                record.pop("gioi_tinh", None)
                supabase.table("nguoi_dung").insert(record).execute()
            thanh_cong += 1
            mat_khau_da_cap.append({"ho_ten": ho_ten, "email": email, "mat_khau": mat_khau_cap})
        except Exception as e:
            loi.append(f"Dong {i}: {str(e)[:120]}")

    # [v2.6] Tra ve danh sach mat khau de GV/QTV phat cho tung nguoi.
    # Day la lan DUY NHAT he thong hien mat khau — sau do chi con ban bam.
    return {
        "thanh_cong": thanh_cong, "loi": loi, "tong": thanh_cong + len(loi),
        "mat_khau_da_cap": mat_khau_da_cap,
        "luu_y": "Moi tai khoan co mat khau rieng. Hay tai ve va phat cho tung nguoi — "
                 "he thong khong hien lai duoc nua.",
    }

class CapNhatHocSinh(BaseModel):
    ho_ten: Optional[str] = None
    email: Optional[str] = None
    email_phu_huynh: Optional[str] = None

@router.put("/hoc-sinh/{id}")
def cap_nhat_hoc_sinh(id: str, body: CapNhatHocSinh, nguoi_dung=Depends(chi_giao_vien)):
    data = {}
    if body.ho_ten is not None: data["ho_ten"] = body.ho_ten
    if body.email is not None: data["email"] = body.email
    if body.email_phu_huynh is not None:
        data["email_phu_huynh"] = body.email_phu_huynh.strip() or None
    if not data:
        raise HTTPException(status_code=400, detail="Khong co du lieu de cap nhat")
    res = supabase.table("nguoi_dung").update(data).eq("id", id).eq("vai_tro", "hoc_sinh").execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Khong tim thay hoc sinh")
    return res.data[0]

def _kiem_tra_quyen_tren_tai_khoan(nguoi_dung: dict, id_muc_tieu: str):
    """[v2.6] Chan leo thang dac quyen.

    Truoc day giao vien co the reset mat khau / vo hieu hoa BAT KY tai khoan nao
    ke ca quan tri vien (chi kiem tra vai tro nguoi goi, khong kiem tra muc tieu)
    -> co the chiem quyen quan tri. Nay rang buoc pham vi ro rang.
    """
    vt = nguoi_dung.get("vai_tro")
    if vt not in ("quan_tri", "giao_vien"):
        raise HTTPException(status_code=403, detail="Khong co quyen")

    muc_tieu = supabase.table("nguoi_dung").select("vai_tro, ten_lop, khoi").eq("id", id_muc_tieu).execute()
    if not muc_tieu.data:
        raise HTTPException(status_code=404, detail="Khong tim thay tai khoan")
    mt = muc_tieu.data[0]

    if vt == "quan_tri":
        return

    # Giao vien: chi thao tac tren HOC SINH thuoc lop minh (hoac khoi neu la truong khoi)
    if mt.get("vai_tro") != "hoc_sinh":
        raise HTTPException(status_code=403, detail="Giao vien chi thao tac duoc tren tai khoan hoc sinh")
    if mt.get("ten_lop") and str(mt["ten_lop"]) == str(nguoi_dung.get("ten_lop")):
        return
    if nguoi_dung.get("la_truong_khoi") and mt.get("khoi") \
       and str(mt["khoi"]) == str(nguoi_dung.get("khoi_phu_trach")):
        return
    raise HTTPException(status_code=403, detail="Chi thao tac duoc tren hoc sinh lop minh phu trach")


@router.put("/{id}/reset-mat-khau")
def reset_mat_khau(id: str, request: Request, nguoi_dung=Depends(lay_nguoi_dung_hien_tai)):
    _kiem_tra_quyen_tren_tai_khoan(nguoi_dung, id)

    # [v2.6] Cap mat khau ngau nhien rieng thay vi mat khau mac dinh dung chung
    mat_khau_moi = tao_mat_khau_ngau_nhien()
    supabase.table("nguoi_dung").update({
        "mat_khau_hash": hash_mat_khau(mat_khau_moi),
        "bat_buoc_doi_mat_khau": True
    }).eq("id", id).execute()

    thu_hoi_phien(id)   # [v2.6] dat lai mat khau -> huy moi phien dang mo cua tai khoan do
    audit.ghi_nhat_ky(audit.DAT_LAI_MAT_KHAU, "Dat lai mat khau cho tai khoan khac",
                      nguoi_dung=nguoi_dung, doi_tuong_id=id, request=request)
    return {
        "message": "Da dat lai mat khau. Hay chuyen mat khau nay cho nguoi dung — he thong khong hien lai.",
        "mat_khau_moi": mat_khau_moi,
    }

@router.delete("/{id}")
def vo_hieu_hoa(id: str, request: Request, nguoi_dung=Depends(lay_nguoi_dung_hien_tai)):
    _kiem_tra_quyen_tren_tai_khoan(nguoi_dung, id)

    supabase.table("nguoi_dung").update({"dang_hoat_dong": False}).eq("id", id).execute()
    thu_hoi_phien(id)   # [v2.6] vo hieu hoa -> dang xuat ngay khoi moi thiet bi
    audit.ghi_nhat_ky(audit.VO_HIEU_HOA_TK, "Vo hieu hoa tai khoan",
                      nguoi_dung=nguoi_dung, doi_tuong_id=id, request=request)
    return {"message": "Da vo hieu hoa tai khoan"}
